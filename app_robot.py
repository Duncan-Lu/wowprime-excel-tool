import os
import streamlit as st
import pandas as pd
import io
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

# 設定網頁標題與排版
st.set_page_config(page_title="服務幫手整合工具 (行動版)", layout="centered")
st.title("🤖 服務幫手整合工具")

# ---- 以下保留你原本的核心數據邏輯 ----

def assign_shift(time):
    morning = datetime.strptime('06:00', '%H:%M').time()
    lunch = datetime.strptime('11:00', '%H:%M').time()
    evening = datetime.strptime('17:00', '%H:%M').time()
    night = datetime.strptime('21:00', '%H:%M').time()
    if morning <= time < lunch:
        return '早班'
    elif lunch <= time < evening:
        return '午班'
    elif evening <= time <= night:
        return '晚班'
    return None

def assign_half_hour_slot(time):
    if pd.isna(time) or not hasattr(time, 'hour'):
        return "無效時間"
    return f"{time.hour:02d}:{'00' if time.minute < 30 else '30'}"

def generate_half_hour_slots(start='06:00', end='21:00'):
    slots, current = [], datetime.strptime(start, '%H:%M')
    end_time = datetime.strptime(end, '%H:%M')
    while current <= end_time:
        slots.append(current.strftime('%H:%M'))
        current += timedelta(minutes=30)
    return slots

def calculate_delivery_counts(df, all_stores, group_by):
    if group_by not in df.columns:
        raise KeyError(f"❌ 缺少必要欄位：{group_by}")
    subtask_cols = ['子任务1餐盘所在层号','子任务2餐盘所在层号','子任务3餐盘所在层号','子任务4餐盘所在层号']
    dest_cols = ['子任务1目的地名称','子任务2目的地名称','子任务3目的地名称','子任务4目的地名称']

    def extract_layers(row):
        result = {}
        for sub, dest in zip(subtask_cols, dest_cols):
            if sub in row and dest in row and pd.notna(row[sub]) and pd.notna(row[dest]):
                d = str(row[dest]).strip()
                layers = [x.strip() for x in str(row[sub]).split(',') if x.strip() != '0']
                result.setdefault(d, set()).update(layers)
        return sum(len(l) for l in result.values())

    df = df.copy()
    df.loc[:, '餐盤總數'] = df.apply(extract_layers, axis=1)
    df.loc[:, '時段'] = df[group_by]
    all_dates = df['日期'].dropna().unique()
    all_slots = generate_half_hour_slots() if group_by == '半小時時段' else df['時段'].dropna().unique()
    idx = pd.MultiIndex.from_product([all_stores, all_dates, all_slots], names=['门店名称','日期','時段'])
    grouped = df.groupby(['门店名称','日期','時段'])['餐盤總數'].sum()
    return grouped.reindex(idx, fill_value=0).reset_index()

def calculate_reception_counts(df, all_stores, group_by):
    if group_by not in df.columns:
        raise KeyError(f"❌ 缺少必要欄位：{group_by}")
    def is_valid_reception_point(pt):
        return bool(str(pt).strip())
    df = df.copy()
    df.loc[:, '目標點總數'] = df['子任务1目的地名称'].astype(str).apply(
        lambda x: len([pt for pt in str(x).split(',') if is_valid_reception_point(pt)]) if pd.notna(x) else 0)
    df.loc[:, '時段'] = df[group_by]
    all_dates = df['日期'].dropna().unique()
    all_slots = generate_half_hour_slots() if group_by == '半小時時段' else df['時段'].dropna().unique()
    idx = pd.MultiIndex.from_product([all_stores, all_dates, all_slots], names=['门店名称','日期','時段'])
    grouped = df.groupby(['门店名称','日期','時段'])['目標點總數'].sum()
    return grouped.reindex(idx, fill_value=0).reset_index()

def mark_abnormal_rows(ws, row_count):
    if row_count <= 0:
        return
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row_idx in range(2, 2 + row_count):
        ws[f"A{row_idx}"].fill = fill

bar_robot_serials = [
    '0C:CF:89:E1:D2:5F', '0C:C6:55:45:FD:EF', '0C:CF:89:9E:19:0C',
    '2C:C3:E6:E8:33:CE', '54:EF:33:CA:3F:76', '54:EF:33:CA:3F:9E',
    '0C:CF:89:E1:16:9D', '0C:CF:89:9E:19:11', '34:7D:E4:9C:04:D9',
    '14:5D:34:EA:56:46', '38:7A:CC:A5:1E:DE', '14:5D:34:EA:55:5D',
    '38:7A:CC:B1:8F:45', '34:7D:E4:9C:09:53', '14:5D:34:EA:55:31',
]

def normalize_na(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.strip()
    return s.replace({'': pd.NA, 'nan': pd.NA, 'NaN': pd.NA, 'None': pd.NA, '-': pd.NA, 'N/A': pd.NA, 'NA': pd.NA})

# ---- 網頁版核心資料處理流程（改為記憶體內處理，不落地讀寫檔案） ----
def process_data_web(uploaded_main_file, uploaded_store_file):
    df = pd.read_excel(uploaded_main_file)
    
    # 讀取選填的「所有門店.xlsx」
    if uploaded_store_file is not None:
        try:
            manual_stores = pd.read_excel(uploaded_store_file, header=None)[0].dropna().unique().tolist()
        except:
            manual_stores = []
    else:
        manual_stores = []
        
    df['门店名称'] = df['门店名称'].astype(str).str.strip()
    actual_stores = df['门店名称'].dropna().unique().tolist()
    all_stores = sorted(set(actual_stores) | set(manual_stores))

    df['任务开始时间'] = pd.to_datetime(df['任务开始时间'], errors='coerce')
    df['日期'] = df['任务开始时间'].dt.date
    df['時間'] = df['任务开始时间'].dt.time
    df['小時'] = df['任务开始时间'].dt.hour
    df['半小時時段'] = df['時間'].apply(assign_half_hour_slot)
    df['班次'] = df['時間'].apply(assign_shift)
    df.dropna(subset=['日期','小時'], inplace=True)

    mode_series = df['任务模式'].astype(str)
    df_delivery  = df[mode_series.str.contains(r'送餐', na=False)].copy()
    df_reception = df[mode_series.str.contains(r'直达|直達', na=False)].copy()

    df_bar_delivery = df_delivery[df_delivery['机器人序列号'].isin(bar_robot_serials)].copy()
    df_bar_reception = df_reception[df_reception['机器人序列号'].isin(bar_robot_serials)].copy()

    # --- 建立 整合報表 記憶體緩衝區 ---
    main_buffer = io.BytesIO()
    with pd.ExcelWriter(main_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='原始數據', index=False)
        df_delivery.to_excel(writer, sheet_name='送餐數據', index=False)
        df_reception.to_excel(writer, sheet_name='收餐數據', index=False)

        target_cols = [
            '子任务1目的地名称','子任务2目的地名称','子任务3目的地名称','子任务4目的地名称',
            '子任务1餐盘所在层号','子任务2餐盘所在层号','子任务3餐盘所在层号','子任务4餐盘所在层号'
        ]
        for col in target_cols:
            if col in df_delivery.columns:
                df_delivery[col] = normalize_na(df_delivery[col])

        if ('子任务1目的地名称' in df_delivery.columns) and ('子任务1餐盘所在层号' in df_delivery.columns):
            severe_mask = df_delivery['子任务1目的地名称'].isna() | df_delivery['子任务1餐盘所在层号'].isna()
        else:
            severe_mask = pd.Series(False, index=df_delivery.index)

        severe_df = df_delivery[severe_mask].copy()

        child1_ok = (
            ('子任务1目的地名称' in df_delivery.columns) and
            ('子任务1餐盘所在层号' in df_delivery.columns) and
            df_delivery['子任务1目的地名称'].notna() & df_delivery['子任务1餐盘所在层号'].notna()
        )
        child2_4_cols = [c for c in target_cols if c.startswith('子任务') and not c.startswith('子任务1') and c in df_delivery.columns]
        if child2_4_cols:
            reminder_mask = child1_ok & df_delivery[child2_4_cols].isna().any(axis=1)
        else:
            reminder_mask = pd.Series(False, index=df_delivery.index)

        reminder_df = df_delivery[reminder_mask & (~severe_mask)].copy()

        def _sort_df(d):
            if ('门店名称' in d.columns) and ('任务开始时间' in d.columns):
                return d.sort_values(['门店名称','任务开始时间'], ascending=[True, True])
            return d

        severe_df = _sort_df(severe_df)
        reminder_df = _sort_df(reminder_df)

        severe_sheet = '異常_嚴重'
        reminder_sheet = '提醒_子任務空白'
        severe_df.to_excel(writer, sheet_name=severe_sheet, index=False)
        ws_severe = writer.sheets[severe_sheet]
        mark_abnormal_rows(ws_severe, len(severe_df))
        reminder_df.to_excel(writer, sheet_name=reminder_sheet, index=False)

        if not reminder_df.empty and ('门店名称' in reminder_df.columns):
            store_sum = reminder_df.groupby('门店名称', dropna=False).size().reset_index(name='提醒筆數')
        else:
            store_sum = pd.DataFrame(columns=['门店名称','提醒筆數'])
        store_sum.to_excel(writer, sheet_name='提醒_門店統計', index=False)

        if not reminder_df.empty and ('日期' in reminder_df.columns) and ('门店名称' in reminder_df.columns):
            daily_sum = reminder_df.groupby(['日期','门店名称'], dropna=False).size().reset_index(name='提醒筆數')
            daily_sum = daily_sum.sort_values(['日期','门店名称'])
        else:
            daily_sum = pd.DataFrame(columns=['日期','门店名称','提醒筆數'])
        daily_sum.to_excel(writer, sheet_name='提醒_日別統計', index=False)

        # 11點前送餐紀錄
        eleven_time = datetime.strptime('11:00', '%H:%M').time()
        before_11_mask = df_delivery['時間'] < eleven_time if '時間' in df_delivery.columns else df_delivery['任务开始时间'].dt.time < eleven_time

        before_11_df = df_delivery[before_11_mask].copy()
        before_11_df = _sort_df(before_11_df)
        before_11_sheet = '異常_11點前送餐'
        before_11_df.to_excel(writer, sheet_name=before_11_sheet, index=False)

        # --- 收集網頁提示警訊資訊 ---
        alert_before_11 = []
        if not before_11_df.empty and ('门店名称' in before_11_df.columns):
            b11_store = before_11_df.groupby('门店名称', dropna=False).size().reset_index(name='11點前筆數')
            alert_before_11 = [f"{row['门店名称']} - {row['11點前筆數']} 筆" for _, row in b11_store.iterrows()]

        half_hour_stats = calculate_delivery_counts(df_delivery, all_stores, '半小時時段')
        half_hour_stats.to_excel(writer, sheet_name='半小時送餐統計', index=False)

        # 忙碌時段為 0 警示
        busy_periods = [
            ('午班', ['12:00', '12:30', '13:00', '13:30']),
            ('晚班', ['17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30'])
        ]
        zero_alerts = []
        for store in all_stores:
            for date in half_hour_stats['日期'].unique():
                for shift_name, slots in busy_periods:
                    sub_df = half_hour_stats[
                        (half_hour_stats['门店名称'] == store) &
                        (half_hour_stats['日期'] == date) &
                        (half_hour_stats['時段'].isin(slots))
                    ]
                    if sub_df['餐盤總數'].sum() == 0:
                        zero_alerts.append(f"⚠️ {store} {date} {shift_name}")

        # 其他常規統計分頁寫入
        calculate_delivery_counts(df_delivery, all_stores, '班次') \
            .pivot_table(index=['门店名称','日期'], columns='時段', values='餐盤總數', fill_value=0) \
            .reset_index().to_excel(writer, sheet_name='班次送餐統計', index=False)

        calculate_reception_counts(df_reception, all_stores, '半小時時段') \
            .to_excel(writer, sheet_name='半小時收餐統計', index=False)

        calculate_reception_counts(df_reception, all_stores, '班次') \
            .pivot_table(index=['门店名称','日期'], columns='時段', values='目標點總數', fill_value=0) \
            .reset_index().to_excel(writer, sheet_name='班次收餐統計', index=False)

        # 返程效率
        delivery_time_column = '送餐总用时（s）'
        return_time_column = '返回总用时（s）'
        if delivery_time_column in df.columns and return_time_column in df.columns:
            df_filtered_delivery = df[(df[delivery_time_column] > 0) & (df[delivery_time_column] <= 300)].copy()
            df_filtered_return = df[(df[return_time_column] > 0) & (df[return_time_column] <= 300)].copy()
            delivery_summary = df_filtered_delivery.groupby(['日期', '门店名称', '班次'])[delivery_time_column].agg(['count', 'mean']).reset_index()
            delivery_summary['mean'] = delivery_summary['mean'].round().astype(int)
            delivery_summary.columns = ['日期', '门店名称', '班次', '送餐总次数', '送餐平均時間（s）']
            return_summary = df_filtered_return.groupby(['日期', '门店名称', '班次'])[return_time_column].agg(['count', 'mean']).reset_index()
            return_summary['mean'] = return_summary['mean'].round().astype(int)
            return_summary.columns = ['日期', '门店名称', '班次', '返回总次数', '返回平均時間（s）']
            all_dates2 = df['日期'].dropna().unique()
            all_shifts2 = df['班次'].dropna().unique()
            all_stores2 = df['门店名称'].dropna().unique()
            full_index = pd.MultiIndex.from_product([all_stores2, all_dates2, all_shifts2], names=['门店名称', '日期', '班次'])
            merged = pd.merge(delivery_summary, return_summary, on=['日期', '门店名称', '班次'], how='outer')
            summary_df = merged.set_index(['门店名称', '日期', '班次']).reindex(full_index, fill_value=0).reset_index()
            df_filtered_delivery.to_excel(writer, sheet_name='送餐符合條件的數據', index=False)
            df_filtered_return.to_excel(writer, sheet_name='返回符合條件的數據', index=False)
            summary_df.to_excel(writer, sheet_name='每日餐期計算結果', index=False)

        # 吧台專屬
        calculate_delivery_counts(df_bar_delivery, all_stores, '半小時時段').to_excel(writer, sheet_name='吧台_半小時送餐統計', index=False)
        吧台班次送餐 = calculate_delivery_counts(df_bar_delivery, all_stores, '班次')
        吧台班次送餐透視表 = 吧台班次送餐.pivot_table(index=['门店名称', '日期'], columns='時段', values='餐盤總數', fill_value=0)
        for col in ['早班', '午班', '晚班']:
            if col not in 吧台班次送餐透視表.columns: 吧台班次送餐透視表[col] = 0
        吧台班次送餐透視表 = 吧台班次送餐透視表[['午班', '晚班', '早班']].reset_index()
        吧台班次送餐透視表.to_excel(writer, sheet_name='吧台_班次送餐統計', index=False)

        calculate_reception_counts(df_bar_reception, all_stores, '半小時時段').to_excel(writer, sheet_name='吧台_半小時收餐統計', index=False)
        吧台班次收餐 = calculate_reception_counts(df_bar_reception, all_stores, '班次')
        吧台班次收餐透視表 = 吧台班次收餐.pivot_table(index=['门店名称', '日期'], columns='時段', values='目標點總數', fill_value=0)
        for col in ['早班', '午班', '晚班']:
            if col not in 吧台班次收餐透視表.columns: 吧台班次收餐透視表[col] = 0
        吧台班次收餐透視表 = 吧台班次收餐透視表[['午班', '晚班', '早班']].reset_index()
        吧台班次收餐透視表.to_excel(writer, sheet_name='吧台_班次收餐統計', index=False)
    
    main_buffer.seek(0)

    # --- 建立 獨立異常報表 記憶體緩衝區 ---
    anomaly_buffer = io.BytesIO()
    with pd.ExcelWriter(anomaly_buffer, engine='openpyxl') as w2:
        severe_df.to_excel(w2, sheet_name='異常_嚴重', index=False)
        reminder_df.to_excel(w2, sheet_name='提醒_子任務空白', index=False)
        before_11_df.to_excel(w2, sheet_name='異常_11點前送餐', index=False)

        store_severe = severe_df.groupby('门店名称').size().reset_index(name='嚴重異常筆數') if not severe_df.empty else pd.DataFrame(columns=['门店名称','嚴重異常筆數'])
        store_reminder = reminder_df.groupby('门店名称').size().reset_index(name='提醒筆數') if not reminder_df.empty else pd.DataFrame(columns=['门店名称','提醒筆數'])
        store_b11 = before_11_df.groupby('门店名称').size().reset_index(name='11點前筆數') if not before_11_df.empty else pd.DataFrame(columns=['门店名称','11點前筆數'])

        if not store_b11.empty:
            store_b11.to_excel(w2, sheet_name='11點前_門店統計', index=False)
        if not store_severe.empty or not store_reminder.empty:
            store_all = pd.merge(store_severe, store_reminder, on='门店名称', how='outer').fillna(0)
            store_all[['嚴重異常筆數','提醒筆數']] = store_all[['嚴重異常筆數','提醒筆數']].astype(int)
        else:
            store_all = pd.DataFrame(columns=['门店名称','嚴重異常筆數','提醒筆數'])
        store_all.to_excel(w2, sheet_name='門店統計', index=False)

        daily_rem = reminder_df.groupby(['日期','门店名称']).size().reset_index(name='提醒筆數').sort_values(['日期','门店名称']) if not reminder_df.empty else pd.DataFrame(columns=['日期','门店名称','提醒筆數'])
        daily_rem.to_excel(w2, sheet_name='提醒_日別統計', index=False)
        
    anomaly_buffer.seek(0)
    
    return main_buffer, anomaly_buffer, alert_before_11, zero_alerts

# ---- 🧱 Streamlit 網頁操作介面 ----

# 1. 檔案上傳區
uploaded_main = st.file_uploader("1. 請上傳 機器人原始數據 Excel (*.xlsx)", type=["xlsx"])
uploaded_stores = st.file_uploader("2. 選擇性上傳 所有門店清單 (所有門店.xlsx，若無可不傳)", type=["xlsx"])

st.write("---")

if uploaded_main:
    if st.button("🚀 開始計算並產生整合報表", type="primary", use_container_width=True):
        with st.spinner("正在進行大數據清洗與餐期計算中..."):
            try:
                main_file, anomaly_file, alert_b11, alert_zero = process_data_web(uploaded_main, uploaded_stores)
                
                st.success("🎉 報表計算完成！請點選下方按鈕下載：")
                
                # 下載按鈕 (手機可以直接儲存至檔案)
                today_str = datetime.today().strftime("%Y%m%d")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="📥 下載 整合數據報表",
                        data=main_file,
                        file_name=f"{today_str}_Duncan數據.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with col2:
                    st.download_button(
                        label="⚠️ 下載 異常提醒專用檔",
                        data=anomaly_file,
                        file_name=f"{today_str}_異常數據.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                # --- 網頁警告訊息區 (取代原本的 messagebox 彈窗) ---
                if alert_b11 or alert_zero:
                    st.write("---")
                    st.subheader("📊 數據重點異常提示")
                    
                    if alert_b11:
                        with st.status("⚠️ 發現 11 點前送餐紀錄", expanded=True):
                            for item in alert_b11:
                                st.write(item)
                                
                    if alert_zero:
                        with st.status("⚠️ 發現 忙碌時段送餐量為 0 之門市", expanded=True):
                            for item in alert_zero:
                                st.write(item)
            except Exception as e:
                st.error(f"計算失敗，請檢查欄位格式是否正確。錯誤訊息: {e}")
